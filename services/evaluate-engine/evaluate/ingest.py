"""Document ingestion — PDF bytes -> per-page text, with an honest quality gate.

Ported by COPY from the bidder engine's app/ingest.py. Not imported: the F13 wall forbids a
shared module between the two products, and `tools/check-wall.sh` fails the build on a stray
`from app…`. Duplication is the cheaper failure here.

Text PDFs are parsed with pypdf. Image-only scans yield almost no text and are REPORTED as
illegible rather than passed through silently — a bidder disqualified because page 14 was a
photograph is the worst outcome this product can produce.
"""

from __future__ import annotations

import csv
import hashlib
import io
import os
import zipfile
from concurrent.futures import ThreadPoolExecutor

from pypdf import PdfReader

from .envelope import ApiError

# A page with almost no extractable text is probably a scan.
_MIN_CHARS_PER_PAGE = 20
# One model call per page; sequential is minutes on a real RFP. Bounded fan-out.
_WORKERS = int(os.environ.get("EVAL_EXTRACT_WORKERS", "8"))
# Below this, a human confirms before the value is trusted.
CONFIRM_THRESHOLD = 0.80


def parse_pdf_pages(data: bytes) -> list[tuple[int, str]]:
    """(page_number, text) per page, 1-indexed."""
    try:
        reader = PdfReader(io.BytesIO(data))
    except Exception as exc:  # noqa: BLE001 — any parse failure is a bad upload
        raise ApiError(400, "BAD_DOCUMENT", f"could not read PDF: {exc}") from exc

    pages: list[tuple[int, str]] = []
    for i, page in enumerate(reader.pages, 1):
        try:
            text = (page.extract_text() or "").strip()
        except Exception:  # noqa: BLE001 — one bad page must not 500 the upload
            text = ""
        pages.append((i, text))
    return pages


def split_legible(pages: list[tuple[int, str]]) -> tuple[list[tuple[int, str]], list[int]]:
    """(legible pages, illegible page numbers). The caller surfaces the latter to the officer."""
    legible = [(p, t) for p, t in pages if len(t) >= _MIN_CHARS_PER_PAGE]
    illegible = [p for p, t in pages if len(t) < _MIN_CHARS_PER_PAGE]
    return legible, illegible


def map_pages(fn, pages: list[tuple[int, str]]) -> list:
    """Fan `fn(page_no, text)` across pages, preserving input order.

    pool.map keeps order and `pages` is page-ascending, so results stay page-ascending — which
    is what keeps extracted anchors in reading order without a re-sort.
    """
    if not pages:
        return []
    with ThreadPoolExecutor(max_workers=_WORKERS) as pool:
        return list(pool.map(lambda pt: fn(pt[0], pt[1]), pages))


# ── bulk intake (F14) ──────────────────────────────────────────────────────────
# A ZIP is untrusted input. Everything below assumes the archive is hostile until proven
# boring: a zip bomb is a denial of service, a `../` entry writes outside the extraction root,
# and a symlink entry can point anywhere on the filesystem. We never write entries to disk at
# all — they are read into memory under a byte cap — but the name checks stay, because the day
# someone adds a temp-file path is the day they matter.

def sha256_of(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def is_archive(filename: str, data: bytes) -> bool:
    return filename.lower().endswith(".zip") or data[:2] == b"PK"


def _safe_entry(name: str) -> bool:
    """Reject traversal, absolute paths and directory entries."""
    if name.endswith("/"):
        return False
    if name.startswith("/") or name.startswith("\\"):
        return False
    parts = name.replace("\\", "/").split("/")
    return ".." not in parts and not any(p.startswith("__MACOSX") for p in parts)


def unpack_archive(data: bytes, *, max_files: int, max_bytes: int) -> list[tuple[str, bytes]]:
    """(filename, bytes) per entry. Raises ApiError on any bound breach — never truncates.

    Silent truncation is the failure mode to avoid here: an officer who drops 600 files and
    sees 500 rows has no way to know which hundred bids are missing, and a missing bid produces
    no error anywhere downstream.
    """
    try:
        zf = zipfile.ZipFile(io.BytesIO(data))
    except zipfile.BadZipFile as exc:
        raise ApiError(400, "BAD_ARCHIVE", f"could not read archive: {exc}") from exc

    entries = [i for i in zf.infolist() if _safe_entry(i.filename)]

    if len(entries) > max_files:
        raise ApiError(413, "ARCHIVE_TOO_LARGE",
                       f"archive holds {len(entries)} files; the limit is {max_files}")
    total = sum(i.file_size for i in entries)
    if total > max_bytes:
        raise ApiError(413, "ARCHIVE_TOO_LARGE",
                       f"archive expands to {total} bytes; the limit is {max_bytes}")

    out: list[tuple[str, bytes]] = []
    for info in entries:
        with zf.open(info) as fh:
            # Read one byte past the declared size: a lying header is the classic bomb.
            payload = fh.read(info.file_size + 1)
        if len(payload) > info.file_size:
            raise ApiError(413, "ARCHIVE_TOO_LARGE",
                           f"entry {info.filename} is larger than its declared size")
        name = info.filename.replace("\\", "/").split("/")[-1]
        if name:
            out.append((name, payload))
    return out


def _display_name(name: str) -> str:
    """A refused entry name, made safe to store, log and render.

    These names are written by the BIDDER, not the officer, and they end up in an audit
    payload, an API response and a screen. Echoing them verbatim cost us a live failure:
    Supabase's WAF sees `../../etc/passwd` in a request body and 403s it, so one crafted entry
    in one bid failed the officer's entire upload with an opaque database error. A bidder
    should not be able to do that.

    The path is what makes the name dangerous and it is also the part with no information in
    it — we already refused the entry, so where it wanted to go does not matter. Keep the leaf,
    say the path was removed.
    """
    leaf = name.replace("\\", "/").rstrip("/").split("/")[-1]
    leaf = leaf.replace("..", "").strip() or "(unnamed)"
    return leaf if leaf == name else f"{leaf} (path removed)"


def rejected_entries(data: bytes) -> list[str]:
    """Entry names an archive contained that we refuse to read (F14-ERR3). Reported, not hidden.

    Names are sanitised for display — see _display_name. Reporting the count and the leaf is
    what the officer needs; reproducing the exact hostile path is what breaks the request.
    """
    try:
        zf = zipfile.ZipFile(io.BytesIO(data))
    except zipfile.BadZipFile:
        return []
    return [_display_name(i.filename) for i in zf.infolist()
            if not _safe_entry(i.filename) and not i.filename.endswith("/")]


# ── format normalisation (F16) ─────────────────────────────────────────────────
# Everything becomes (page_number, text) — the shape the rest of this engine already consumes.
# A spreadsheet's "page" is a sheet and its anchors are cell references; that is a deliberate
# stretch of the word, and it is cheaper than teaching four downstream modules a second shape.

def _xlsx_pages(data: bytes) -> list[tuple[int, str]]:
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — any openpyxl failure is a bad upload
        raise ApiError(400, "BAD_DOCUMENT", f"could not read spreadsheet: {exc}") from exc

    pages: list[tuple[int, str]] = []
    for n, ws in enumerate(wb.worksheets, 1):
        lines = [f"SHEET: {ws.title}"]
        for row in ws.iter_rows():
            cells = [f"{c.coordinate}={c.value}" for c in row if c.value not in (None, "")]
            if cells:
                lines.append(" | ".join(cells))
        pages.append((n, "\n".join(lines)))
    wb.close()
    return pages


def _csv_pages(data: bytes) -> list[tuple[int, str]]:
    text = data.decode("utf-8", errors="replace")
    rows = list(csv.reader(io.StringIO(text)))
    body = "\n".join(" | ".join(c for c in r if c) for r in rows if any(r))
    return [(1, body)]


def parse_pages(filename: str, data: bytes) -> list[tuple[int, str]]:
    """Dispatch on format. Unsupported types raise rather than returning an empty document —
    an empty parse looks exactly like a bid with nothing in it."""
    lower = filename.lower()
    if lower.endswith(".pdf") or data[:5] == b"%PDF-":
        return parse_pdf_pages(data)
    if lower.endswith((".xlsx", ".xlsm")):
        return _xlsx_pages(data)
    if lower.endswith(".csv"):
        return _csv_pages(data)
    raise ApiError(422, "UNSUPPORTED_FORMAT",
                   f"{filename}: only PDF, XLSX and CSV are read in this version")


def render_page_png(data: bytes, page_no: int, scale: float = 2.0) -> bytes:
    """One PDF page as a PNG, for the OCR fallback (F16).

    pypdfium2 rather than PyMuPDF: PyMuPDF is AGPL, which is the wrong licence for software
    sold to a public authority. pypdfium2 is permissive and needs no system binary.
    """
    import pypdfium2 as pdfium

    pdf = pdfium.PdfDocument(data)
    try:
        page = pdf[page_no - 1]
        image = page.render(scale=scale).to_pil()
        buf = io.BytesIO()
        image.save(buf, format="PNG")
        return buf.getvalue()
    finally:
        pdf.close()
