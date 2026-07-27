"""XLSX round-trip (G-AC4/G-AC5).

The scenario these protect: a bid manager exports, an SME edits on a plane, the file comes
back. Everything the user changed must land, and nothing that came from the locked tender
document may change.
"""

from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook

from app.deterministic.matrix import MatrixRow, MatrixRowStatus, plan_import
from app.deterministic.types import RequirementLevel, SourceAnchor
from app.envelope import ApiError
from app.matrix_export import KEY_HEADER, build_xlsx, parse_xlsx

M = RequirementLevel.MANDATORY


def _row(cid: str = "c1", **kw) -> MatrixRow:
    base = dict(
        criterion_id=cid,
        requirement_text="The bidder shall submit a CA-certified turnover certificate.",
        requirement_level=M,
        anchor=SourceAnchor(page=12, clause="4.1(a)"),
        evidence_required="CA certificate",
    )
    return MatrixRow(**{**base, **kw})


def _edit(data: bytes, header: str, value, row: int = 2) -> bytes:
    """Edit a cell by HEADER NAME, the way a user edits by column heading."""
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    headers = {c.value: c.column for c in ws[1]}
    ws.cell(row=row, column=headers[header]).value = value
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- export -----------------------------------------------------------------------------


def test_export_writes_one_line_per_row_with_the_key_column_hidden():
    ws = load_workbook(io.BytesIO(build_xlsx([_row("c1"), _row("c2")], "GEM/2026/B/5127401"))).active

    assert ws["A1"].value == KEY_HEADER
    assert ws.column_dimensions["A"].hidden is True  # the user never sees a uuid...
    assert [ws.cell(row=r, column=1).value for r in (2, 3)] == ["c1", "c2"]  # ...but it is there
    assert ws.freeze_panes == "B2"


def test_export_of_an_empty_matrix_is_a_valid_sheet_with_only_headers():
    ws = load_workbook(io.BytesIO(build_xlsx([], "t"))).active
    assert ws["A1"].value == KEY_HEADER
    assert ws["A2"].value is None


def test_export_survives_a_row_with_no_anchor():
    ws = load_workbook(io.BytesIO(build_xlsx([_row(anchor=None)], "t"))).active
    assert ws["B2"].value is None and ws["C2"].value is None


# --- round-trip fidelity (G-AC4) ---------------------------------------------------------


def test_export_then_import_with_no_edits_changes_nothing():
    rows = [_row("c1"), _row("c2", requirement_level=RequirementLevel.DESIRABLE, anchor=None)]
    parsed, conflicts = parse_xlsx(build_xlsx(rows, "t"))

    assert conflicts == ()
    plan = plan_import(rows, parsed, ignore_fields=("owner",))
    assert plan.ok is True
    assert plan.updates == ()
    assert plan.unchanged == 2


def test_the_trailing_note_rows_are_not_parsed_as_requirements():
    # The sheet carries an instructions line and a tender caption below the data.
    parsed, _ = parse_xlsx(build_xlsx([_row()], "GEM/2026/B/5127401"))
    assert len(parsed) == 1


@pytest.mark.parametrize(
    ("header", "value", "attr"),
    [
        ("Response ref", "§3.2", "response_ref"),
        ("Status", "drafted", "status"),
        ("Due date", "2026-08-01", "due_date"),
        ("Notes", "ask finance for the FY25 figure", "notes"),
    ],
)
def test_an_offline_edit_to_an_editable_column_lands(header, value, attr):
    existing = [_row()]
    edited = _edit(build_xlsx(existing, "t"), header, value)

    parsed, conflicts = parse_xlsx(edited)
    plan = plan_import(existing, parsed, ignore_fields=("owner",))

    assert conflicts == () and plan.ok is True
    assert len(plan.updates) == 1
    got = getattr(plan.updates[0], attr)
    assert (got.value if attr == "status" else got) == value


@pytest.mark.parametrize(
    ("header", "value"),
    [
        ("Requirement", "a requirement the bidder would prefer"),
        ("Level", "desirable"),
        ("Page", 99),
        ("Clause", "9.9(z)"),
    ],
)
def test_editing_a_locked_column_offline_is_a_conflict(header, value):
    existing = [_row()]
    parsed, _ = parse_xlsx(_edit(build_xlsx(existing, "t"), header, value))

    plan = plan_import(existing, parsed, ignore_fields=("owner",))
    assert plan.ok is False
    assert plan.updates == ()


def test_owner_is_displayed_but_not_written_back():
    # The sheet shows a name; the column stores a user id. Resolving one to the other is a
    # guess, so the import reports it as ignored instead of guessing.
    existing = [_row(owner="user-1")]
    parsed, _ = parse_xlsx(_edit(build_xlsx(existing, "t"), "Owner", "someone else"))

    plan = plan_import(existing, parsed, ignore_fields=("owner",))
    assert plan.ok is True
    assert plan.updates == () and plan.unchanged == 1


# --- a sheet that has been through a real user ------------------------------------------


def test_headers_are_read_by_name_so_an_inserted_column_does_not_shift_the_data():
    # Inserting a column is the first thing a user does. A positional parser would silently
    # read the wrong column and write plausible garbage into a compliance artifact.
    wb = load_workbook(io.BytesIO(build_xlsx([_row()], "t")))
    ws = wb.active
    ws.insert_cols(2)
    ws["B1"] = "My own working column"
    ws["B2"] = "scratch"
    buf = io.BytesIO()
    wb.save(buf)

    parsed, conflicts = parse_xlsx(buf.getvalue())
    assert conflicts == ()
    assert parsed[0].requirement_text.startswith("The bidder shall submit")


def test_a_reordered_sheet_still_matches_rows_to_requirements():
    rows = [_row("c1"), _row("c2", requirement_text="Bidders must hold ISO 9001.")]
    wb = load_workbook(io.BytesIO(build_xlsx(rows, "t")))
    ws = wb.active
    ws["A2"], ws["A3"] = "c2", "c1"          # user sorted the sheet
    ws["D2"], ws["D3"] = ws["D3"].value, ws["D2"].value
    buf = io.BytesIO()
    wb.save(buf)

    parsed, _ = parse_xlsx(buf.getvalue())
    plan = plan_import(rows, parsed, ignore_fields=("owner",))
    assert plan.ok is True and plan.unchanged == 2


def test_a_blank_line_left_in_the_sheet_is_skipped():
    wb = load_workbook(io.BytesIO(build_xlsx([_row()], "t")))
    ws = wb.active
    ws.insert_rows(2)
    buf = io.BytesIO()
    wb.save(buf)
    parsed, _ = parse_xlsx(buf.getvalue())
    assert len(parsed) == 1


def test_a_typo_in_status_is_a_conflict_not_a_silent_reset():
    # Coercing an unrecognised status to 'not_started' would move a finished requirement back
    # to the top of someone's queue without telling them.
    parsed, conflicts = parse_xlsx(_edit(build_xlsx([_row()], "t"), "Status", "Drafted!!"))

    assert parsed == ()
    assert len(conflicts) == 1
    assert conflicts[0].field == "status"
    assert "not_started" in conflicts[0].reason


def test_status_matching_is_case_insensitive():
    parsed, conflicts = parse_xlsx(_edit(build_xlsx([_row()], "t"), "Status", "Approved"))
    assert conflicts == ()
    assert parsed[0].status is MatrixRowStatus.APPROVED


def test_an_emptied_status_cell_reads_as_not_started():
    parsed, conflicts = parse_xlsx(_edit(build_xlsx([_row()], "t"), "Status", None))
    assert conflicts == () and parsed[0].status is MatrixRowStatus.NOT_STARTED


def test_an_unrecognised_level_is_a_conflict():
    parsed, conflicts = parse_xlsx(_edit(build_xlsx([_row()], "t"), "Level", "critical"))
    assert parsed == () and conflicts[0].field == "requirement_level"


def test_a_non_numeric_page_degrades_to_no_anchor_rather_than_crashing():
    parsed, conflicts = parse_xlsx(_edit(build_xlsx([_row()], "t"), "Page", "twelve"))
    assert conflicts == () and parsed[0].anchor is None


# --- structural damage -------------------------------------------------------------------


def test_deleting_the_key_column_fails_loudly():
    # The alternative — matching on requirement text — breaks the moment someone fixes a typo,
    # and would rewrite the wrong requirement while reporting success.
    wb = load_workbook(io.BytesIO(build_xlsx([_row()], "t")))
    wb.active.delete_cols(1)
    buf = io.BytesIO()
    wb.save(buf)

    with pytest.raises(ApiError) as exc:
        parse_xlsx(buf.getvalue())
    assert exc.value.code == "MATRIX_IMPORT_CONFLICT"
    assert KEY_HEADER in exc.value.message


def test_a_file_that_is_not_a_workbook_is_a_400_not_a_500():
    with pytest.raises(ApiError) as exc:
        parse_xlsx(b"this is a PDF, actually")
    assert exc.value.status == 400
