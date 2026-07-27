"""XLSX round-trip for the compliance matrix (G-FR4).

Bid desks live in Excel. A product that refuses to meet them there loses to a spreadsheet,
so the matrix exports, gets edited on a plane, and comes back — without losing traceability.

Pure functions over rows: no DB, no HTTP, so this unit-tests without a live anything (same
shape as docx_export.py).

Two rules the parser exists to enforce:

  - **Row identity travels in a hidden key column.** Matching on requirement text would break
    the moment someone fixed a typo, and matching on row order would break the moment someone
    sorted the sheet — which is the first thing anyone does in Excel.
  - **Headers are read by NAME, not position.** Users insert columns. A positional parser
    silently reads the wrong column and writes plausible garbage into a compliance artifact.
"""

from __future__ import annotations

import io
from collections.abc import Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .deterministic.matrix import ImportConflict, MatrixRow, MatrixRowStatus
from .deterministic.types import RequirementLevel, SourceAnchor
from .envelope import ApiError

KEY_HEADER = "criterion_id"

#: (header, width, editable-in-Excel). Order is the sheet's column order.
COLUMNS: tuple[tuple[str, int, bool], ...] = (
    (KEY_HEADER, 12, False),
    ("Page", 7, False),
    ("Clause", 14, False),
    ("Requirement", 80, False),
    ("Level", 16, False),
    ("Evidence required", 32, False),
    ("Response ref", 18, True),
    ("Owner", 22, False),  # displayed for information; ignored on import (see plan_import)
    ("Status", 14, True),
    ("Due date", 12, True),
    ("Notes", 40, True),
)

_HEADER_FILL = PatternFill("solid", fgColor="F2F2F7")   # design tokens: surface-alt
_LOCKED_FILL = PatternFill("solid", fgColor="FAFAFC")
_STATUS_VALUES = tuple(s.value for s in MatrixRowStatus)


def build_xlsx(rows: Sequence[MatrixRow], tender_title: str) -> bytes:
    """Render the matrix as a workbook. The key column is hidden, not absent."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Compliance Matrix"

    ws.append([h for h, _w, _e in COLUMNS])
    for i, (_h, width, editable) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
        cell = ws.cell(row=1, column=i)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        if not editable:
            cell.font = Font(bold=True, italic=True)

    for row in rows:
        ws.append(
            [
                row.criterion_id,
                row.anchor.page if row.anchor else None,
                row.anchor.clause if row.anchor else None,
                row.requirement_text,
                row.requirement_level.value,
                row.evidence_required,
                row.response_ref,
                row.owner,
                row.status.value,
                row.due_date,
                row.notes,
            ]
        )

    # The key column is hidden rather than removed: the user never sees a uuid, and the sheet
    # still knows which requirement each line is. Deleting the column breaks import loudly
    # (missing key header) rather than silently matching on text.
    ws.column_dimensions["A"].hidden = True
    ws.freeze_panes = "B2"

    last = ws.max_row
    for r in range(2, last + 1):
        ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="top")
        for c in (2, 3, 4, 5, 6):
            ws.cell(row=r, column=c).fill = _LOCKED_FILL

    note = ws.cell(row=last + 2, column=2)
    note.value = (
        "Shaded columns come from the locked tender document and are not editable — "
        "changes to them are reported as conflicts on import. Owner is set in the app. "
        f"Valid Status values: {', '.join(_STATUS_VALUES)}."
    )
    note.font = Font(italic=True, size=9)

    ws.cell(row=last + 4, column=2).value = f"Tender: {tender_title}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell_str(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def parse_xlsx(data: bytes) -> tuple[tuple[MatrixRow, ...], tuple[ImportConflict, ...]]:
    """Read an uploaded sheet back into rows.

    Structural damage (not a workbook, key column deleted) raises — there is nothing to
    reconcile. Bad VALUES inside otherwise-sound rows come back as conflicts, so the user sees
    every problem at once instead of one per upload.
    """
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True)
    except Exception as exc:  # noqa: BLE001 — any load failure is a bad upload, surfaced as 400
        raise ApiError(400, "MATRIX_IMPORT_CONFLICT", f"could not read workbook: {exc}") from exc

    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = {_cell_str(h): i for i, h in enumerate(header_row) if _cell_str(h)}

    if KEY_HEADER not in headers:
        raise ApiError(
            400,
            "MATRIX_IMPORT_CONFLICT",
            f"the '{KEY_HEADER}' column is missing — re-export the matrix and edit that file; "
            "without it a row cannot be matched to its requirement",
        )

    def get(row: tuple, header: str) -> str | None:
        idx = headers.get(header)
        return _cell_str(row[idx]) if idx is not None and idx < len(row) else None

    rows: list[MatrixRow] = []
    conflicts: list[ImportConflict] = []

    for raw in ws.iter_rows(min_row=2, values_only=True):
        criterion_id = get(raw, KEY_HEADER)
        if not criterion_id:
            continue  # the trailing note rows, and any blank line a user left behind

        status_text = (get(raw, "Status") or MatrixRowStatus.NOT_STARTED.value).lower()
        try:
            status = MatrixRowStatus(status_text)
        except ValueError:
            # Never coerce: a typo'd status silently becoming 'not_started' would move a
            # finished requirement back to the top of someone's queue.
            conflicts.append(
                ImportConflict(
                    criterion_id=criterion_id,
                    field="status",
                    existing="",
                    incoming=status_text,
                    reason=f"status must be one of: {', '.join(_STATUS_VALUES)}",
                )
            )
            continue

        level_text = (get(raw, "Level") or "").lower()
        try:
            level = RequirementLevel(level_text)
        except ValueError:
            conflicts.append(
                ImportConflict(
                    criterion_id=criterion_id,
                    field="requirement_level",
                    existing="",
                    incoming=level_text,
                    reason="requirement level is not a recognised value",
                )
            )
            continue

        page = get(raw, "Page")
        clause = get(raw, "Clause")
        has_page = bool(page) and page.isdigit()
        anchor = SourceAnchor(page=int(page), clause=clause or "") if has_page else None

        rows.append(
            MatrixRow(
                criterion_id=criterion_id,
                requirement_text=get(raw, "Requirement") or "",
                requirement_level=level,
                anchor=anchor,
                evidence_required=get(raw, "Evidence required"),
                response_ref=get(raw, "Response ref"),
                owner=get(raw, "Owner"),
                status=status,
                due_date=get(raw, "Due date"),
                notes=get(raw, "Notes"),
            )
        )

    return tuple(rows), tuple(conflicts)
