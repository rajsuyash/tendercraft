"""Ingestion — tender document -> per-page text -> extracted criteria (Module A).

Text-based PDFs are parsed with pypdf. Scanned PDFs (image-only pages) yield little text
and route to manual review (EC-1); OCR of scans is a later step. Extraction runs per page
so every criterion keeps its page anchor (A-AC3).

A tender arrives as a PACKAGE, not a file: an NIT, a handful of annexures, and a BOQ
spreadsheet, each of which can carry eligibility clauses. They ingest as ONE tender, because
three separate tenders with three readiness checklists is not what the buyer published. That
makes the page number alone an unresolvable anchor — "p.4" of which document? — so every page
carries the label of the document it came from, and criteria keep the LOCAL page a human can
actually turn to rather than a running count across the package.
"""

from __future__ import annotations

import csv
import io
import logging
import os
import zipfile
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass

from pypdf import PdfReader

from .deterministic.boq import BoqRow, rows_to_line_items
from .deterministic.shred import unmapped_sentences
from .deterministic.tender_meta import extract_tender_meta
from .deterministic.types import EXTRACTION_CONFIRM_THRESHOLD
from .envelope import ApiError

log = logging.getLogger("tendercraft.ingest")

# A page with almost no extractable text is probably a scan — flag for manual OCR (EC-1).
_MIN_CHARS_PER_PAGE = 20
# Per-page extraction is one Gemini call each; sequential is minutes on a real RFP.
# Fan out across pages, bounded so we don't hammer the API. Order restored after.
_EXTRACT_WORKERS = int(os.environ.get("INGEST_EXTRACT_WORKERS", "8"))
# Office formats are zips. A 2 MB xlsx can expand to gigabytes (knowledge.py hit this first).
_MAX_UNCOMPRESSED = 100 * 1024 * 1024
_SPREADSHEET_EXTS = ("xlsx", "xlsm")


@dataclass(frozen=True)
class SourcePage:
    """One readable unit of the package, and where a human would go to find it.

    `document` is the label shown beside the anchor — a filename, or "file · sheet" for a
    spreadsheet, where the sheet name is the locator and the index alone would mean nothing.
    `page` is local to that document: the page of the PDF, or the sheet's position.
    """

    document: str
    page: int
    text: str


def parse_pdf_pages(data: bytes) -> list[tuple[int, str]]:
    """Return (page_number, text) for each page. 1-indexed pages."""
    try:
        reader = PdfReader(io.BytesIO(data))
    except Exception as exc:  # noqa: BLE001 — any parse failure is a bad upload, surfaced as 400
        raise ApiError(400, "BAD_DOCUMENT", f"could not read PDF: {exc}") from exc

    pages: list[tuple[int, str]] = []
    for i, page in enumerate(reader.pages, 1):
        try:
            text = (page.extract_text() or "").strip()
        except Exception:  # noqa: BLE001 — a single unreadable page shouldn't 500; treat as illegible
            text = ""
        pages.append((i, text))
    return pages


def parse_spreadsheet_pages(filename: str, data: bytes) -> list[SourcePage]:
    """One page per worksheet. BOQs and eligibility matrices live in sheets, not prose.

    Cells are joined per row so the extractor reads a requirement table as lines rather than
    a wall of values. Formulas are read as their last computed value (`data_only`) — a bid
    desk's "=SUM(...)" is worth nothing to a reader who cannot recalculate it.
    """
    _guard_zip_bomb(data)
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — any parse failure is a bad upload (400)
        raise ApiError(400, "BAD_DOCUMENT", f"could not read spreadsheet: {exc}") from exc

    pages: list[SourcePage] = []
    try:
        for index, sheet in enumerate(wb.worksheets, 1):
            lines = [
                " | ".join(str(v).strip() for v in row if v is not None and str(v).strip())
                for row in sheet.iter_rows(values_only=True)
            ]
            pages.append(
                SourcePage(f"{filename} · {sheet.title}", index, "\n".join(x for x in lines if x))
            )
    finally:
        wb.close()  # read_only leaves file handles open otherwise
    return pages


def parse_boq_rows(filename: str, data: bytes) -> list[BoqRow]:
    """Recover a spreadsheet's schedule of items as ROWS, additively.

    `parse_spreadsheet_pages` above is untouched and still produces one page of joined text per
    worksheet — criteria extraction and the unmapped-sentence denominator never notice this
    function exists. That separation is deliberate: schedule parsing is new and unproven, and it
    must not be able to change what the TOM contains.

    Only the openpyxl read lives here; the column mapping is pure and testable in
    app/deterministic/boq.py. A sheet with no recognisable header contributes nothing rather
    than a guess.
    """
    _guard_zip_bomb(data)
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — any parse failure is a bad upload (400)
        raise ApiError(400, "BAD_DOCUMENT", f"could not read spreadsheet: {exc}") from exc

    items: list[BoqRow] = []
    try:
        for index, sheet in enumerate(wb.worksheets, 1):
            rows = list(sheet.iter_rows(values_only=True))
            items.extend(
                rows_to_line_items(f"{filename} · {sheet.title}", index, rows)
            )
    finally:
        wb.close()  # read_only leaves file handles open otherwise
    return items


def parse_package_boq(documents: list[tuple[str, bytes]]) -> list[BoqRow]:
    """Every schedule line in the package. A BOQ failure is never fatal to an upload —
    criteria extraction is the product; the schedule is an addition to it."""
    found: list[BoqRow] = []
    for filename, data in documents:
        if filename.lower().rsplit(".", 1)[-1] not in _SPREADSHEET_EXTS:
            continue
        try:
            found.extend(parse_boq_rows(filename, data))
        except ApiError:
            log.warning("BOQ parse failed for %s — schedule skipped, ingest continues", filename)
    return found


def parse_csv_pages(filename: str, data: bytes) -> list[SourcePage]:
    text = data.decode("utf-8", errors="replace")
    rows = [
        " | ".join(c.strip() for c in row if c.strip())
        for row in csv.reader(io.StringIO(text))
    ]
    return [SourcePage(filename, 1, "\n".join(r for r in rows if r))]


def parse_document_pages(filename: str, data: bytes) -> list[SourcePage]:
    """Dispatch on extension. Unknown formats are rejected, never silently read as bytes."""
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if ext == "pdf":
        return [SourcePage(filename, page, text) for page, text in parse_pdf_pages(data)]
    if ext in _SPREADSHEET_EXTS:
        return parse_spreadsheet_pages(filename, data)
    if ext == "csv":
        return parse_csv_pages(filename, data)
    raise ApiError(
        400, "UNSUPPORTED_FORMAT",
        f"{filename}: only PDF, XLSX, XLSM and CSV can be read as tender documents",
    )


def _guard_zip_bomb(data: bytes) -> None:
    """Reject office files whose uncompressed size would blow up memory."""
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as z:
            if sum(i.file_size for i in z.infolist()) > _MAX_UNCOMPRESSED:
                raise ApiError(413, "FILE_TOO_LARGE", "document expands to an unsafe size")
    except zipfile.BadZipFile:
        pass  # not a zip — the format parser rejects it with a better message


def number_package(
    documents: list[SourcePage],
) -> tuple[list[tuple[int, str]], dict[int, SourcePage]]:
    """Give the package one running page sequence, and keep the way back.

    The extractor needs a page number that is unique across the whole package (two documents
    both have a page 3), while the criterion a human reads must name the document and its own
    page. So: number globally for extraction, map back before anything is persisted.
    """
    pages = [(i, doc.text) for i, doc in enumerate(documents, 1)]
    return pages, {i: doc for i, doc in enumerate(documents, 1)}


def ingest_pages(pages: list[tuple[int, str]]) -> dict:
    """Run the extractor over pages; return criteria rows + ingestion stats.

    Imported lazily so unit tests can monkeypatch the extractor without a live key.
    """
    from pipeline.extractor import extract_from_page

    illegible_pages = [p for p, t in pages if len(t) < _MIN_CHARS_PER_PAGE]
    legible = [(p, t) for p, t in pages if len(t) >= _MIN_CHARS_PER_PAGE]

    # One Gemini call per page, fanned out. pool.map preserves input order and `legible` is
    # page-ascending, so rows stay page-ascending without a re-sort (anchors ordered, A-AC3).
    with ThreadPoolExecutor(max_workers=_EXTRACT_WORKERS) as pool:
        per_page = pool.map(lambda pt: extract_from_page(pt[1], pt[0]), legible)

    rows: list[dict] = []
    low_conf = 0
    for criteria in per_page:
        for c in criteria:
            if c.needs_confirmation:
                low_conf += 1
            rows.append(
                {
                    "verbatim_text": c.verbatim_text,
                    "category": c.category,
                    "requirement_level": c.requirement_level,
                    "confidence": c.confidence,
                    "confirmed": False,
                    "anchor_page": c.anchor_page,
                    "anchor_clause": c.anchor_clause or None,
                    "evidence_required": c.evidence_required or None,
                    "evaluation_weight": c.evaluation_weight,
                }
            )

    # The document states its own identity on page one. Reading it is why a bid stops
    # being called "rfp.pdf" on every screen.
    meta = extract_tender_meta([text for _page, text in pages])

    # G-FR2, and it has to happen HERE. This codebase does not persist tender text — pages are
    # parsed in memory and discarded — so this is the only moment the denominator can be
    # computed without asking the user to upload the document a second time.
    unmapped = unmapped_sentences(
        legible, [(r["anchor_page"], r["verbatim_text"]) for r in rows]
    )

    return {
        "meta": meta,
        "criteria_rows": rows,
        "unmapped_rows": [{"sentence": u.text, "page": u.page} for u in unmapped],
        "extracted": len(rows),
        "low_confidence": low_conf,
        "illegible_pages": illegible_pages,
        "confirm_threshold": EXTRACTION_CONFIRM_THRESHOLD,
    }
